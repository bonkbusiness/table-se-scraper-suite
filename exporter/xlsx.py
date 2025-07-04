"""
exporter/xlsx.py

Exports product dictionaries to XLSX (Excel), with QC pipeline integration, parent/subcategory coloring
(using category color logic), improved logging, and enhanced layout/styling.

Features:
- Exports a list of product dicts to an XLSX file, sorted by a configurable key ("Namn" by default).
- Uses a fixed column order for consistency with Table.se exports.
- Includes "Kategori (parent)" and "Kategori (sub)" columns for parent and subcategory information.
- Applies unique pastel colors to category/subcategory cells using logic from scraper.utils.
- Improved styling: header freeze, alternating row banding, autofilter, wrapped text, etc.
- Uses logging for status and error messages (integrates with scraper.logging).
- Compatible with QC pipeline (via export_products_with_qc).

Datapoints/columns exported (see scraper/product.py extraction):
    - Namn
    - Artikelnummer
    - Färg
    - Material
    - Serie
    - Pris exkl. moms (värde)
    - Pris exkl. moms (enhet)
    - Pris inkl. moms (värde)
    - Pris inkl. moms (enhet)
    - Längd (värde)
    - Längd (enhet)
    - Bredd (värde)
    - Bredd (enhet)
    - Höjd (värde)
    - Höjd (enhet)
    - Djup (värde)
    - Djup (enhet)
    - Diameter (värde)
    - Diameter (enhet)
    - Kapacitet (värde)
    - Kapacitet (enhet)
    - Volym (värde)
    - Volym (enhet)
    - Vikt (värde)
    - Vikt (enhet)
    - Data (text)
    - Kategori (parent)
    - Kategori (sub)
    - Produktbild-URL
    - Produkt-URL
    - Beskrivning
    - Extra data

API:
- export_to_xlsx(data, filename, sort_key="Namn")
- export_products_with_qc(products, filename, error_filename=None)
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from scraper.logging import get_logger
from scraper.utils import build_category_colors, make_output_filename

logger = get_logger("xlsx-export")

PRODUCT_COLUMN_ORDER = [
    "Namn",
    "Artikelnummer",
    "Färg",
    "Material",
    "Serie",
    "Pris exkl. moms (värde)",
    "Pris exkl. moms (enhet)",
    "Pris inkl. moms (värde)",
    "Pris inkl. moms (enhet)",
    "Längd (värde)", "Längd (enhet)",
    "Bredd (värde)", "Bredd (enhet)",
    "Höjd (värde)", "Höjd (enhet)",
    "Djup (värde)", "Djup (enhet)",
    "Diameter (värde)", "Diameter (enhet)",
    "Kapacitet (värde)", "Kapacitet (enhet)",
    "Volym (värde)", "Volym (enhet)",
    "Vikt (värde)", "Vikt (enhet)",
    "Data (text)",
    "Kategori (parent)",
    "Kategori (sub)",
    "Produktbild-URL",
    "Produkt-URL",
    "Beskrivning",
    "Extra data",
]

def to_argb(color):
    color = str(color).lstrip('#')
    if len(color) == 6:
        return "FF" + color.upper()
    elif len(color) == 8:
        return color.upper()
    else:
        return "FFFFFFFF"

def export_to_xlsx(data, filename=None, sort_key="Namn"):
    """
    Export a list of product dicts to XLSX, sorted by sort_key.
    Each product dict may include all fields listed in PRODUCT_COLUMN_ORDER.
    Returns the filename or None on error.
    """
    if not data:
        logger.warning("Ingen data att exportera till XLSX.")
        return None
    if filename is None:
        filename = make_output_filename('products', 'xlsx', 'export')
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    try:
        data_sorted = sorted(data, key=lambda x: x.get(sort_key, "").lower())
        wb = Workbook()
        ws = wb.active
        ws.title = "Produkter"

        get_color = build_category_colors(data_sorted)

        # Header row: bold white, dark bg, freeze, autofilter
        for col_num, col in enumerate(PRODUCT_COLUMN_ORDER, 1):
            cell = ws.cell(row=1, column=col_num, value=col)
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor="FF212121")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="medium", color="FFB0BEC5"))
        ws.freeze_panes = ws["A2"]
        ws.auto_filter.ref = ws.dimensions

        # Data rows
        band_color = PatternFill("solid", fgColor="FFF5F5F5")
        for row_num, row in enumerate(data_sorted, 2):
            is_band = (row_num % 2 == 0)
            for col_num, col in enumerate(PRODUCT_COLUMN_ORDER, 1):
                value = row.get(col, "")
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if is_band:
                    cell.fill = band_color
                if col == "Kategori (parent)" or col == "Kategori (sub)":
                    color = get_color(row)
                    if color:
                        color = to_argb(color)
                        if color != "FFFFFFFF":
                            cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if col in ("Produktbild-URL", "Produkt-URL") and value:
                    cell.hyperlink = value
                    cell.style = "Hyperlink"
                if any(kw in col for kw in ("värde", "Pris", "Längd", "Bredd", "Höjd", "Djup", "Diameter", "Kapacitet", "Volym")):
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                elif col not in ("Produktbild-URL", "Produkt-URL"):
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = Border(left=Side(style="thin", color="FFD3D3D3"),
                                    right=Side(style="thin", color="FFD3D3D3"),
                                    top=Side(style="thin", color="FFD3D3D3"),
                                    bottom=Side(style="thin", color="FFD3D3D3"))
        for col_num, col in enumerate(PRODUCT_COLUMN_ORDER, 1):
            max_length = max(
                [len(str(row.get(col, ""))) for row in data_sorted] + [len(col)]
            )
            ws.column_dimensions[get_column_letter(col_num)].width = min(max(12, max_length + 2), 50)
        wb.save(filename)
        logger.info(f"Export till XLSX klar: {filename}")
        return filename
    except Exception as e:
        logger.error(f"Fel vid sparande av XLSX: {e}")
        return None

def export_products_with_qc(products, filename=None, error_filename=None):
    """
    Main entrypoint for the QC pipeline: deduplicate, check completeness, and export to XLSX.
    Optionally export products with missing fields to a separate XLSX file.
    """
    from exporter.qc import deduplicate_products, check_field_completeness, export_errors_to_xlsx

    deduped = deduplicate_products(products)
    incomplete = check_field_completeness(deduped)
    valid = [p for p in deduped if p not in incomplete]
    exported = export_to_xlsx(valid, filename)
    logger.info(f"QC-pipeline: Exporterade {len(valid)} produkter till {exported}")
    if (error_filename or incomplete):
        if error_filename is None:
            error_filename = make_output_filename('errors', 'xlsx', 'error')
        export_errors_to_xlsx(
            [{"error_type": "missing_fields", "product": p} for p in incomplete],
            error_filename
        )
        logger.info(f"QC-pipeline: Exporterade {len(incomplete)} felaktiga produkter till {error_filename}")
    return exported