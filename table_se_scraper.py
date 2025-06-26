# ========================
# 1. Install dependencies
# ========================
# !pip install requests beautifulsoup4 openpyxl

# ========================
# 2. Imports and Logging
# ========================
import requests
from bs4 import BeautifulSoup
import re
import unicodedata
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import urljoin, urlparse
import logging
import os
import csv
import traceback

from exclusions import EXCLUDED_CATEGORIES, EXCLUDED_PRODUCTS
from product_cache import get_cached_product, update_cache, hash_content
from table_se_scraper_backend_enhanced import main_enhanced
from table_se_scraper_performance import setup_logging, robust_scrape
from table_se_smart_scanner import smart_scan_products

setup_logging()

def logprint(msg):
    print(msg)
    logging.info(msg)

def should_skip(category_name):
    #return category_name in EXCLUDED_CATEGORIES
    return False


BASE_URL = "https://www.table.se"

# ========================
# 3. Utility functions
# ========================
def normalize_text(text):
    if not text:
        return ""
    text = text.lower()
    trans = str.maketrans("åäö", "aao")
    text = text.translate(trans)
    text = unicodedata.normalize('NFKD', text).encode('ascii','ignore').decode()
    return text.strip()
    
def get_soup(url, timeout=20):
    """
    Downloads a URL and returns a BeautifulSoup object (or None if failed).
    """
    try:
        resp = requests.get(url, timeout=timeout)
        resp.raise_for_status()
        return BeautifulSoup(resp.text, "html.parser")
    except Exception as e:
        logprint(f"Fel vid hämtning av {url}: {e}")
        return None

def extract_only_number_value(text):
    """
    Extracts a numeric value (as string, with dot as decimal separator) from a messy string.
    Example: "1 234,00 kr" -> "1234.00"
    """
    import re
    if not text:
        return ""
    # Remove all non-digit, non-decimal-separator characters
    # Accept both , and . as decimal separator
    # Remove spaces (thousand separators)
    cleaned = text.replace(" ", "").replace("\xa0", "")
    # Replace comma with dot for decimals
    cleaned = cleaned.replace(",", ".")
    # Extract the first number in the string
    match = re.search(r"\d*\.?\d+", cleaned)
    return match.group(0) if match else ""
    
def extract_only_numbers(text):
    """Extract only digits from the input string."""
    return "".join(filter(str.isdigit, str(text)))

def get_all_headers(data):
    """
    Returns a list of all unique keys appearing in any dict in data, prioritizing main fields first.
    """
    if not data:
        return []
    headers = set()
    for row in data:
        headers.update(row.keys())
    priority = ["Category", "Subcategory", "Sub-Subcategory"]
    ordered = [h for h in priority if h in headers]
    rest = sorted(h for h in headers if h not in priority)
    return ordered + rest

def backup_export_to_csv(data, filename=None, base_name="table_produkter_backup"):
    """
    Fallback to CSV export if XLSX fails.
    """
    if not data:
        print("Ingen data att exportera till CSV.")
        return None
    if filename is None:
        export_dir = "/content" if os.path.exists("/content") else "."
        filename = os.path.join(export_dir, f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    headers = set()
    for row in data:
        headers.update(row.keys())
    headers = sorted(headers)
    try:
        with open(filename, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for row in data:
                writer.writerow(row)
        print(f"Backup export till CSV klar: {filename}")
        return filename
    except Exception as e:
        print(f"Fel vid backup-CSV-export: {e}")
        logging.error(f"CSV backup export failed: {e}")
        return None

def export_to_xlsx(data, base_name="table_produkter"):
    if not data:
        print("Ingen data att exportera till XLSX.")
        return None
    export_dir = "/content" if os.path.exists("/content") else "."
    filename = os.path.join(export_dir, f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Produkter"

    headers = get_all_headers(data)
    ws.append(headers)

    # Formatting headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF212121")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color="FFB0BEC5"))

    for row in data:
        ws.append([row.get(h, "") for h in headers])
        row_idx = ws.max_row
        category = row.get("Category") or row.get("category") or ""
        subcategory = row.get("Subcategory") or row.get("subcategory") or ""
        pastel_color = pastel_color_for_category(subcategory) if 'pastel_color_for_category' in globals() and pastel_color_for_category(subcategory) != "FFF5F5F5" else (pastel_color_for_category(category) if 'pastel_color_for_category' in globals() else "FFFFFFFF")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", fgColor=pastel_color)
            cell.font = Font(color="FF212121")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin", color="FFCFD8DC"),
                right=Side(style="thin", color="FFCFD8DC"),
                top=Side(style="thin", color="FFCFD8DC"),
                bottom=Side(style="thin", color="FFCFD8DC"),
            )

    for col in ws.columns:
        max_length = max(len(str(cell.value) or "") for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 6

    try:
        wb.save(filename)
        print(f"Export till XLSX klar: {filename}")
    except Exception as e:
        print(f"Fel vid sparande av XLSX: {e}")
        logging.error(f"XLSX export failed: {e}")
        return None
    return filename

def safe_export_to_xlsx_with_colab_backup(data, base_name="table_produkter"):
    """
    Tries to export to XLSX, falls back to CSV if it fails.
    Handles Colab environment gracefully and attempts download if in Colab.
    Returns the filename and a flag indicating if fallback was used.
    """
    export_dir = "/content" if os.path.exists("/content") else "."
    filename = os.path.join(export_dir, f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    try:
        result = export_to_xlsx(data, base_name)
        if result and os.path.exists(result):
            try:
                from google.colab import files
                files.download(result)
            except ImportError:
                pass
            return result, False
        else:
            raise Exception("XLSX export failed, result file missing.")
    except Exception as e:
        print(f"XLSX-export misslyckades: {e}\nFörsöker backup till CSV...")
        tb = traceback.format_exc()
        logging.error(f"XLSX export failed with traceback:\n{tb}")
        backup_filename = os.path.join(export_dir, f"{base_name}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        backup_result = backup_export_to_csv(data, backup_filename)
        if backup_result and os.path.exists(backup_result):
            try:
                from google.colab import files
                files.download(backup_result)
            except ImportError:
                pass
            print("Varning: Resultat exporterades som CSV istället för XLSX på grund av problem.")
            return backup_result, True
        else:
            print("Backup-export misslyckades helt!")
            return None, True

def parse_measurements(text):
    """
    Tries to extract measurements (mått) from messy Swedish product info text.
    Returns a dict with keys like:
      - "Mått (text)"
      - "Längd (värde)", "Längd (enhet)"
      - "Bredd (värde)", "Bredd (enhet)"
      - "Höjd (värde)", "Höjd (enhet)"
      - "Diameter (värde)", "Diameter (enhet)"
    """
    result = {"Mått (text)": text}
    if not text:
        return result

    # Try to find label:value pairs
    patterns = [
        (r"Längd\s*[:=]?\s*([\d,.]+)\s*(cm|mm|m)?", "Längd"),
        (r"Bredd\s*[:=]?\s*([\d,.]+)\s*(cm|mm|m)?", "Bredd"),
        (r"Höjd\s*[:=]?\s*([\d,.]+)\s*(cm|mm|m)?", "Höjd"),
        (r"Diameter\s*[:=]?\s*([\d,.]+)\s*(cm|mm|m)?", "Diameter"),
    ]
    for pat, label in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            val = m.group(1).replace(",", ".")
            unit = m.group(2) if m.group(2) else ""
            result[f"{label} (värde)"] = val
            result[f"{label} (enhet)"] = unit

    # Try generic "mått" or 120x60x75 cm etc
    m = re.search(
        r"(?<!\d)(\d{2,5}(?:[.,]\d+)?)\s*[x×*\/]\s*(\d{2,5}(?:[.,]\d+)?)\s*[x×*\/]\s*(\d{2,5}(?:[.,]\d+)?)(?:\s*(cm|mm|m))?",
        text, re.IGNORECASE)
    if m:
        vals = [m.group(i).replace(",", ".") for i in range(1, 4)]
        unit = m.group(4) if m.group(4) else ""
        result["Längd (värde)"] = vals[0]
        result["Längd (enhet)"] = unit
        result["Bredd (värde)"] = vals[1]
        result["Bredd (enhet)"] = unit
        result["Höjd (värde)"] = vals[2]
        result["Höjd (enhet)"] = unit

    # Also handle "Diameter 30 cm" or "Ø30cm"
    m = re.search(r"(?:diameter|ø)\s*[:=]?\s*([\d,.]+)\s*(cm|mm|m)?", text, re.IGNORECASE)
    if m:
        val = m.group(1).replace(",", ".")
        unit = m.group(2) if m.group(2) else ""
        result["Diameter (värde)"] = val
        result["Diameter (enhet)"] = unit

    # If nothing was extracted except Mått (text), log for manual review
    if len(result) == 1:
        logging.info(f"parse_measurements: Could not parse measurements from: '{text}'")

    return result

# ========================
# 4. Category extraction (3 levels deep)
# ========================
def extract_category_tree():
    resp = requests.get(BASE_URL + "/produkter/")
    soup = BeautifulSoup(resp.text, "html.parser")
    main_categories = []
    seen = set()
    for a in soup.find_all("a", href=True):
        href = a['href']
        url = urljoin(BASE_URL, href)
        parsed = urlparse(href)
        if parsed.path.startswith("/produkter/") and not "/page/" in parsed.path and not "/nyheter/" in parsed.path:
            path_parts = [p for p in parsed.path.split("/") if p]
            if len(path_parts) == 2 or (len(path_parts) == 3 and path_parts[2] == ""):
                catname = a.get_text(strip=True)
                if catname and catname != "HEM" and url not in seen:
                    seen.add(url)
                    main_categories.append({"name": catname, "url": url})
    logprint(f"Hittade {len(main_categories)} huvudkategorier")

    tree = []
    for cat in main_categories:
        node = {"name": cat["name"], "url": cat["url"], "subs": []}
        sub_soup = get_soup(cat["url"])
        subcats = []
        seen_sub = set()
        if sub_soup:
            for a in sub_soup.find_all("a", href=True):
                href = a['href']
                url_sub = urljoin(BASE_URL, href)
                parsed = urlparse(href)
                path_parts = [p for p in parsed.path.split("/") if p]
                if (
                    len(path_parts) == 3 and
                    path_parts[0] == "produkter" and
                    path_parts[1] == urlparse(cat["url"]).path.split("/")[2]
                ):
                    catname = a.get_text(strip=True)
                    if catname and catname != "HEM" and url_sub not in seen_sub:
                        seen_sub.add(url_sub)
                        subcats.append({"name": catname, "url": url_sub})
        # Go one level deeper (sub-subcategories)
        for sub in subcats:
            subsub_soup = get_soup(sub["url"])
            subsubs = []
            seen_subsub = set()
            if subsub_soup:
                for a in subsub_soup.find_all("a", href=True):
                    href = a['href']
                    url2 = urljoin(BASE_URL, href)
                    parsed2 = urlparse(href)
                    path_parts2 = [p for p in parsed2.path.split("/") if p]
                    if (
                        len(path_parts2) == 4 and
                        path_parts2[0] == "produkter" and
                        path_parts2[1] == urlparse(cat["url"]).path.split("/")[2] and
                        path_parts2[2] == urlparse(sub["url"]).path.split("/")[3]
                    ):
                        name2 = a.get_text(strip=True)
                        if name2 and name2 != "HEM" and url2 not in seen_subsub:
                            seen_subsub.add(url2)
                            subsubs.append({"name": name2, "url": url2})
            sub["subs"] = subsubs
        node["subs"] = subcats
        tree.append(node)
    return tree

# ========================
# 5. Scraper functions (3-level deep)
# ========================

#import re

# PATCH: Robust parse_value_unit for scraping
def parse_value_unit(text):
    """
    Splits a string like '12 cm' or '10,5L' into ('12', 'cm') or ('10.5', 'L').
    Returns ('', '') if nothing found or input is None/empty.
    """
    if not text:
        return "", ""
    text = str(text).replace(",", ".")
    match = re.search(r"([\d.]+)\s*([a-zA-ZåäöÅÄÖ%]*)", text)
    if match:
        value, unit = match.group(1), match.group(2)
        return value.strip(), unit.strip()
    return "", ""

def parse_measurements(matt_text):
    """
    Parses measurement text and returns a dict of measurement values.
    Finds Längd, Bredd, Höjd, Djup (or D), Diameter.
    """
    result = {}
    if not matt_text:
        return result
    lines = matt_text.split(",")
    for line in lines:
        parts = line.strip().split()
        if not parts:
            continue
        label = parts[0].capitalize()
        value, unit = "", ""
        if len(parts) > 1:
            value, unit = parse_value_unit(" ".join(parts[1:]))
        # Standardize keys
        if label in ["Längd", "L"]:
            result["Längd (värde)"] = value
            result["Längd (enhet)"] = unit
        elif label in ["Bredd", "B"]:
            result["Bredd (värde)"] = value
            result["Bredd (enhet)"] = unit
        elif label in ["Höjd", "H"]:
            result["Höjd (värde)"] = value
            result["Höjd (enhet)"] = unit
        elif label in ["Djup", "D"]:
            result["Djup (värde)"] = value
            result["Djup (enhet)"] = unit
        elif label in ["Diameter", "Diam." ,"Diam"]:
            result["Diameter (värde)"] = value
            result["Diameter (enhet)"] = unit
        else:
            # fallback for Mått (text)
            result["Mått (text)"] = matt_text
    return result

def extract_product_data(product_url):
    logging.info(f"Extracting: {product_url}")
    soup = get_soup(product_url)
    if not soup:
        logging.warning(f"Soup is None for {product_url}")
        return None

    # Artikelnummer: <strong> inside .woocommerce-product-details__short-description
    short_desc = soup.select_one(".woocommerce-product-details__short-description")
    artikelnummer = ""
    if short_desc:
        strong = short_desc.find("strong")
        if strong:
            artikelnummer = strong.get_text(strip=True)
    artikelnummer = extract_only_numbers(artikelnummer)

    content_hash = hash_content(soup.prettify())

    # Try the cache
    cached = get_cached_product(artikelnummer, content_hash)
    if cached:
        logprint(f"Produkt {artikelnummer} laddad från cache.")
        return cached

    namn = ""
    selectors = [
        "h1.edgtf-single-product-title[itemprop='name']",
        "h1.product_title"
    ]
    for sel in selectors:
        el = soup.select_one(sel)
        if el:
            namn = el.get_text(strip=True)
            break

    pris_inkl_elem = soup.select_one(".product_price_in")
    pris_inkl = (
        pris_inkl_elem.get_text(strip=True)
        if pris_inkl_elem else ""
    )
    pris_inkl = extract_only_number_value(pris_inkl)

    pris_exkl_elem = soup.select_one(".product_price_ex")
    pris_exkl = (
        pris_exkl_elem.get_text(strip=True)
        if pris_exkl_elem else ""
    )
    pris_exkl = extract_only_number_value(pris_exkl)

    produktbild_url = ""
    img = soup.select_one(".woocommerce-product-gallery__image img")
    if img and img.get("src"):
        produktbild_url = img.get("src")

    more_info = soup.select_one('.product_more_info.vc_col-md-6')
    info_dict = {}
    if more_info:
        for p in more_info.find_all('p'):
            lines = []
            for elem in p.contents:
                if isinstance(elem, str):
                    lines.extend(elem.split('\n'))
                elif getattr(elem, 'name', None) == 'br':
                    lines.append('\n')
                else:
                    lines.append(elem.get_text())
            text = ''.join(lines)
            for line in text.split('\n'):
                line = line.strip()
                if not line or ':' not in line:
                    continue
                label, value = line.split(':', 1)
                label = label.strip().capitalize()
                value = value.strip()
                info_dict[label] = value

    farg = info_dict.get('Färg', '')
    material = info_dict.get('Material', '')
    serie = info_dict.get('Serie', '')
    matt_text = info_dict.get('Mått', '') or info_dict.get('Mått (text)', '')
    diameter_text = info_dict.get('Diameter', '')
    kapacitet_text = info_dict.get('Kapacitet', '')
    volym_text = info_dict.get('Volym', '')
    djup_text = info_dict.get('Djup', '') or info_dict.get('D', '')

    mått_dict = parse_measurements(matt_text)
    diameter_v, diameter_e = parse_value_unit(diameter_text)
    if not diameter_v and mått_dict.get("Diameter (värde)"):
        diameter_v = mått_dict.get("Diameter (värde)")
        diameter_e = mått_dict.get("Diameter (enhet)")
    kap_v, kap_e = parse_value_unit(kapacitet_text)
    vol_v, vol_e = parse_value_unit(volym_text)
    längd_v, längd_e = mått_dict.get("Längd (värde)"), mått_dict.get("Längd (enhet)")
    bredd_v, bredd_e = mått_dict.get("Bredd (värde)"), mått_dict.get("Bredd (enhet)")
    höjd_v, höjd_e = mått_dict.get("Höjd (värde)"), mått_dict.get("Höjd (enhet)")
    # Djup: Prefer explicit info_dict, then mått_dict
    djup_v, djup_e = "", ""
    if djup_text:
        djup_v, djup_e = parse_value_unit(djup_text)
    if not djup_v and mått_dict.get("Djup (värde)"):
        djup_v = mått_dict.get("Djup (värde)")
        djup_e = mått_dict.get("Djup (enhet)")

    data = {
        "Namn": namn,
        "Artikelnummer": artikelnummer,
        "Färg": farg,
        "Material": material,
        "Serie": serie,
        "Pris exkl. moms (värde)": pris_exkl,
        "Pris exkl. moms (enhet)": "kr" if pris_exkl else "",
        "Pris inkl. moms (värde)": pris_inkl,
        "Pris inkl. moms (enhet)": "kr" if pris_inkl else "",
        "Mått (text)": mått_dict.get("Mått (text)", matt_text),
        "Längd (värde)": längd_v,
        "Längd (enhet)": längd_e,
        "Bredd (värde)": bredd_v,
        "Bredd (enhet)": bredd_e,
        "Höjd (värde)": höjd_v,
        "Höjd (enhet)": höjd_e,
        "Djup (värde)": djup_v,
        "Djup (enhet)": djup_e,
        "Diameter (värde)": diameter_v,
        "Diameter (enhet)": diameter_e,
        "Kapacitet (värde)": kap_v,
        "Kapacitet (enhet)": kap_e,
        "Volym (värde)": vol_v,
        "Volym (enhet)": vol_e,
        "Produktbild-URL": produktbild_url,
        "Produkt-URL": product_url
    }

    update_cache(artikelnummer, data, content_hash)
    logprint(f"Extraherad produkt: {namn} (URL: {product_url})")
    return data

def export_to_xlsx(data, base_name="export"):
    """
    Export a list of product dicts to XLSX with explicit column order.
    Returns the filename.
    """
    COLUMN_ORDER = [
        "Namn",
        "Artikelnummer",
        "Färg",
        "Material",
        "Serie",
        "Pris exkl. moms (värde)",
        "Pris exkl. moms (enhet)",
        "Pris inkl. moms (värde)",
        "Pris inkl. moms (enhet)",
        "Mått (text)",
        "Längd (värde)", "Längd (enhet)",
        "Bredd (värde)", "Bredd (enhet)",
        "Höjd (värde)", "Höjd (enhet)",
        "Djup (värde)", "Djup (enhet)",
        "Diameter (värde)", "Diameter (enhet)",
        "Kapacitet (värde)", "Kapacitet (enhet)",
        "Volym (värde)", "Volym (enhet)",
        "Produktbild-URL",
        "Produkt-URL"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Produkter"

    # Write header row
    for col_num, col in enumerate(COLUMN_ORDER, 1):
        ws.cell(row=1, column=col_num, value=col)

    # Write data rows
    for row_num, row in enumerate(data, 2):
        for col_num, col in enumerate(COLUMN_ORDER, 1):
            ws.cell(row=row_num, column=col_num, value=row.get(col, ""))

    # Auto-width columns
    for col_num, col in enumerate(COLUMN_ORDER, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = max(12, len(col) + 2)

    # Save
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{base_name}_{now}.xlsx"
    wb.save(filename)
    return filename

def backup_export_to_csv(data, base_name="export_backup"):
    """
    Export a list of product dicts to CSV with explicit column order.
    Returns the filename.
    """
    COLUMN_ORDER = [
        "Namn",
        "Artikelnummer",
        "Färg",
        "Material",
        "Serie",
        "Pris exkl. moms (värde)",
        "Pris exkl. moms (enhet)",
        "Pris inkl. moms (värde)",
        "Pris inkl. moms (enhet)",
        "Mått (text)",
        "Längd (värde)", "Längd (enhet)",
        "Bredd (värde)", "Bredd (enhet)",
        "Höjd (värde)", "Höjd (enhet)",
        "Djup (värde)", "Djup (enhet)",
        "Diameter (värde)", "Diameter (enhet)",
        "Kapacitet (värde)", "Kapacitet (enhet)",
        "Volym (värde)", "Volym (enhet)",
        "Produktbild-URL",
        "Produkt-URL"
    ]
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{base_name}_{now}.csv"
    with open(filename, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMN_ORDER)
        writer.writeheader()
        for row in data:
            writer.writerow({col: row.get(col, "") for col in COLUMN_ORDER})
    return filename

    # ===================
    # More Info Section
    # ===================
    more_info = soup.select_one('.product_more_info.vc_col-md-6')
    info_dict = {}
    if more_info:
        for p in more_info.find_all('p'):
            # Process <p> content with <br> for each line
            lines = []
            for elem in p.contents:
                if isinstance(elem, str):
                    lines.extend(elem.split('\n'))
                elif getattr(elem, 'name', None) == 'br':
                    lines.append('\n')
                else:
                    lines.append(elem.get_text())
            text = ''.join(lines)
            for line in text.split('\n'):
                line = line.strip()
                if not line or ':' not in line:
                    continue
                label, value = line.split(':', 1)
                label = label.strip().capitalize()
                value = value.strip()
                info_dict[label] = value

    farg = info_dict.get('Färg', '')
    material = info_dict.get('Material', '')
    serie = info_dict.get('Serie', '')
    matt_text = info_dict.get('Mått', '') or info_dict.get('Mått (text)', '')
    diameter_text = info_dict.get('Diameter', '')
    kapacitet_text = info_dict.get('Kapacitet', '')
    volym_text = info_dict.get('Volym', '')

    # PATCH: Parse measurements with new logic (L->Längd, etc., no negatives, cm normalization)
    mått_dict = parse_measurements(matt_text)
    diameter_v, diameter_e = parse_value_unit(diameter_text)
    if not diameter_v and mått_dict.get("Diameter (värde)"):
        diameter_v = mått_dict.get("Diameter (värde)")
        diameter_e = mått_dict.get("Diameter (enhet)")
    kap_v, kap_e = parse_value_unit(kapacitet_text)
    vol_v, vol_e = parse_value_unit(volym_text)
    längd_v, längd_e = mått_dict.get("Längd (värde)"), mått_dict.get("Längd (enhet)")
    bredd_v, bredd_e = mått_dict.get("Bredd (värde)"), mått_dict.get("Bredd (enhet)")
    höjd_v, höjd_e = mått_dict.get("Höjd (värde)"), mått_dict.get("Höjd (enhet)")

    data = {
        "Namn": namn,
        "Artikelnummer": artikelnummer,
        "Pris exkl. moms (värde)": pris_exkl,
        "Pris exkl. moms (enhet)": "kr" if pris_exkl else "",
        "Pris inkl. moms (värde)": pris_inkl,
        "Pris inkl. moms (enhet)": "kr" if pris_inkl else "",
        "Mått (text)": mått_dict.get("Mått (text)"),
        "Längd (värde)": längd_v,
        "Längd (enhet)": längd_e,
        "Bredd (värde)": bredd_v,
        "Bredd (enhet)": bredd_e,
        "Höjd (värde)": höjd_v,
        "Höjd (enhet)": höjd_e,
        "Diameter (värde)": diameter_v,
        "Diameter (enhet)": diameter_e,
        "Kapacitet (värde)": kap_v,
        "Kapacitet (enhet)": kap_e,
        "Volym (värde)": vol_v,
        "Volym (enhet)": vol_e,
        "Färg": farg,
        "Material": material,
        "Serie": serie,
        "Produktbild-URL": produktbild_url,
        "Produkt-URL": product_url
    }

    # Update cache after successful extraction
    update_cache(artikelnummer, data, content_hash)
    logprint(f"Extraherad produkt: {namn} (URL: {product_url})")
    return data

def extract_products_from_category(category_url):
    soup = get_soup(category_url)
    if not soup:
        return []
    product_urls = []
    # Page through if there are multiple pages
    next_page = True
    page_number = 1
    while next_page:
        prods = soup.select(".product a.woocommerce-LoopProduct-link, .products .product a")
        for a in prods:
            href = a.get("href")
            if href:
                url = urljoin(BASE_URL, href)
                if url not in product_urls:
                    product_urls.append(url)
        # Look for next page
        next_btn = soup.select_one(".page-numbers .next")
        if next_btn and next_btn.get("href"):
            next_url = urljoin(BASE_URL, next_btn.get("href"))
            page_number += 1
            soup = get_soup(next_url)
            if not soup:
                break
        else:
            next_page = False
    return product_urls

# ========================
# 6. XLSX Exporter
# ========================
def pastel_color_for_category(category):
    pastel_palette = {
        "Möbler":             "FFB3E5FC", # Light Blue
        "Dukning":            "FFFFF9C4", # Light Yellow
        "Tält":               "FFDCEDC8", # Light Green
        "Bar & barutrustning":"FFFFCCBC", # Light Orange
        "Servering":          "FFFFF8E1", # Light Cream
        "Köksutrustning":     "FFD1C4E9", # Light Purple
        "Engångsartiklar":    "FFFFFDE7", # Pale Yellow
        "Garderob & entré":   "FFE1BEE7", # Pastel Lavender
        "Teknik & scen":      "FFB2EBF2", # Pastel Cyan
        "Containrar":         "FFFFE0B2", # Pastel Peach
        "Hyra container":     "FFFFF3E0", # Pastel Apricot
        "Köpa container":     "FFF8BBD0", # Pastel Pink
        "Self storage":       "FFB2DFDB", # Pastel Mint
        "Kylcontainrar":      "FFC5CAE9", # Pastel Indigo
        "Förrådscontainrar":  "FFFFF9C4", # Light Yellow
        "Kök & disk":         "FFDCEDC8", # Pastel Green
        "Toalettbodar":       "FFFFFDE7", # Pale Yellow
        "Kontorsbodar":       "FFE1F5FE", # Light Aquamarine
        "Eventcontainrar":    "FFFFF8E1", # Cream
        "Specialcontainrar":  "FFF3E5F5", # Pastel Purple
        "Containertillbehör": "FFF0F4C3", # Pastel Lime
        "Begagnade containrar":"FFD7CCC8", # Pastel Brown
        "Flytt & förvaringsservice":"FFFFF9C4", # Light Yellow
        "Transporter":        "FFE0F7FA", # Pastel Blue
        "Bord":               "FFE3F2FD",
        "Stolar":             "FFE0F7FA",
        "Soffor & fåtöljer":  "FFF8BBD0",
        "Konferensmöbler":    "FFDCEDC8",
        "Loungemöbler":       "FFD7CCC8",
        "Utemöbler":          "FFFFF9C4",
        "Övriga möbler":      "FFFCE4EC",
        "Glas":               "FFE1F5FE",
        "Porslin":            "FFF3E5F5",
        "Bestick":            "FFEFEBE9",
        "Bordsdekor":         "FFF8BBD0",
        "Linne & överdrag":   "FFE0F7FA",
        "Tälttillbehör":      "FFF1F8E9",
        # Add more as needed
    }
    return pastel_palette.get(category, "FFF5F5F5")
    
def get_all_headers(data):
    """
    Returns a list of all unique keys appearing in any dict in data, prioritizing main fields first.
    """
    if not data:
        return []
    headers = set()
    for row in data:
        headers.update(row.keys())
    priority = ["Category", "Subcategory", "Sub-Subcategory"]
    ordered = [h for h in priority if h in headers]
    rest = sorted(h for h in headers if h not in priority)
    return ordered + rest

def backup_export_to_csv(data, base_name="table_produkter_backup"):
    """
    Fallback to CSV export if XLSX fails.
    """
    if not data:
        print("Ingen data att exportera till CSV.")
        return None
    filename = f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    headers = set()
    for row in data:
        headers.update(row.keys())
    headers = sorted(headers)
    try:
        with open(filename, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for row in data:
                writer.writerow(row)
        print(f"Backup export till CSV klar: {filename}")
        return filename
    except Exception as e:
        print(f"Fel vid backup-CSV-export: {e}")
        logging.error(f"CSV backup export failed: {e}")
        return None

def export_to_xlsx(data, base_name="table_produkter"):
    logging.info(f"DEBUG: export_to_xlsx called with data type: {type(data)}")
    if isinstance(data, list):
        logging.info(f"DEBUG: Data length: {len(data)}")
        if len(data) > 0:
            logging.info(f"DEBUG: First row: {data[0]}")
    else:
        logging.info(f"DEBUG: Data value: {data}")
    if not data:
        print("Ingen data att exportera till XLSX.")
        return None
    filename = f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Produkter"

    headers = get_all_headers(data)
    ws.append(headers)

    # Formatting headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF212121")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color="FFB0BEC5"))

    for row in data:
        ws.append([row.get(h, "") for h in headers])
        row_idx = ws.max_row
        category = row.get("Category") or row.get("category") or ""
        subcategory = row.get("Subcategory") or row.get("subcategory") or ""
        # pastel_color_for_category must be defined elsewhere in your codebase
        pastel_color = pastel_color_for_category(subcategory) if pastel_color_for_category(subcategory) != "FFF5F5F5" else pastel_color_for_category(category)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", fgColor=pastel_color)
            cell.font = Font(color="FF212121")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin", color="FFCFD8DC"),
                right=Side(style="thin", color="FFCFD8DC"),
                top=Side(style="thin", color="FFCFD8DC"),
                bottom=Side(style="thin", color="FFCFD8DC"),
            )

    for col in ws.columns:
        max_length = max(len(str(cell.value) or "") for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 6

    try:
        wb.save(filename)
        print(f"Export till XLSX klar: {filename}")
    except Exception as e:
        print(f"Fel vid sparande av XLSX: {e}")
        logging.error(f"XLSX export failed: {e}")
        return None
    return filename

def safe_export_to_xlsx_with_colab_backup(data, base_name="table_produkter"):
    """
    Tries to export to XLSX, falls back to CSV if it fails.
    Handles Colab environment gracefully and attempts download if in Colab.
    Returns the filename and a flag indicating if fallback was used.
    """
    # Always export to /content in Colab
    export_dir = "/content" if os.path.exists("/content") else "."
    filename = os.path.join(export_dir, f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    try:
        result = export_to_xlsx(data, base_name)
        if result and os.path.exists(result):
            try:
                from google.colab import files
                files.download(result)
            except ImportError:
                pass
            return result, False
        else:
            raise Exception("XLSX export failed, result file missing.")
    except Exception as e:
        print(f"XLSX-export misslyckades: {e}\nFörsöker backup till CSV...")
        backup_filename = os.path.join(export_dir, f"{base_name}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        backup_result = backup_export_to_csv(data, backup_filename)
        if backup_result and os.path.exists(backup_result):
            try:
                from google.colab import files
                files.download(backup_result)
            except ImportError:
                pass
            print("Varning: Resultat exporterades som CSV istället för XLSX på grund av problem.")
            return backup_result, True
        else:
            print("Backup-export misslyckades helt!")
            return None, True

def safe_export_to_xlsx_with_backup(data, base_name="table_produkter"):
    """
    Tries to export to XLSX, falls back to CSV if it fails.
    Returns the filename and a flag indicating if fallback was used.
    """
    filename = export_to_xlsx(data, base_name)
    if filename is not None:
        return filename, False
    print("XLSX-export misslyckades, försöker backup till CSV...")
    backup_filename = backup_export_to_csv(data, base_name + "_backup")
    return backup_filename, True if backup_filename else False

def export_errors_to_xlsx(errors, base_name="table_produkter_errors"):
    if not errors:
        print("Inga valideringsfel att exportera.")
        return None
    filename = f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Produktfel"
    ws.append(["Index", "Feltyp", "Produktinfo"])
    for idx, err in enumerate(errors):
        ws.append([
            idx + 1,
            err.get("error_type", str(err.get("type", ""))),
            str(err.get("product", err))
        ])
    try:
        wb.save(filename)
        print(f"Export av fel till XLSX klar: {filename}")
    except Exception as e:
        print(f"Fel vid sparande av fel-XLSX: {e}")
        logging.error(f"XLSX error export failed: {e}")
        # Backup to CSV for errors too
        backup_filename = backup_export_to_csv(errors, base_name + "_backup")
        if backup_filename:
            print("Felrapport exporterades som CSV istället för XLSX.")
            return backup_filename
        return None
    return filename

# ========================
# 7. Enhanced Main Entrypoint (Parallelized, Smart Scan, Separate Error XLSX)
# ========================
def enhanced_main_with_scan_and_error_file():
    exported_file, fallback_used, error_traceback = main_enhanced(
        extract_category_tree_func=extract_category_tree,
        skip_func=should_skip,
        extract_func=extract_product_data,
        export_func=export_to_xlsx,                  # main export
        max_workers=8,
        fallback_export_func=backup_export_to_csv    # fallback
    )

    # --- ADD DEBUGGING HERE ---
    # Check what was returned from scraping before export
    if isinstance(exported_file, list):
        logging.warning(f"DEBUG: 'exported_file' is a list with {len(exported_file)} entries.")
        if len(exported_file) > 0:
            logging.warning(f"DEBUG: First product: {exported_file[0]}")
        else:
            logging.warning("DEBUG: 'exported_file' is an EMPTY list.")
    elif exported_file is None:
        logging.warning("DEBUG: 'exported_file' is None (scraping/export failed or returned nothing).")
    else:
        logging.warning(f"DEBUG: 'exported_file' type: {type(exported_file)}. Value: {exported_file}")

    if exported_file is None or (isinstance(exported_file, list) and len(exported_file) == 0):
        logprint("Ingen data skrapades eller exporten misslyckades helt.")
        if error_traceback:
            print("FEL OCH TRACEBACK:\n", error_traceback)
        return None, None

    # If the exported_file is a data list, log details
    if isinstance(exported_file, list):
        logging.info(f"DEBUG: Exported data is a list with {len(exported_file)} entries.")
        if len(exported_file) > 0:
            logging.info(f"DEBUG: First entry: {exported_file[0]}")
        else:
            logging.warning("DEBUG: Exported data list is empty.")
    else:
        logging.info(f"DEBUG: Exported data type is {type(exported_file)}. Value: {exported_file}")

    # If main_enhanced does not do smart scan, do it here:
    try:
        scanned_products, product_errors = smart_scan_products([])  # default empty
        # If exported_file is a list (products), scan them
        if isinstance(exported_file, list):
            scanned_products, product_errors = smart_scan_products(exported_file)
        elif isinstance(exported_file, str):
            # Assume exported_file is the filename, skip scanning
            scanned_products, product_errors = [], []
    except Exception as e:
        print("Fel vid smart scanning av produkter:", e)
        scanned_products, product_errors = [], []

    error_xlsx = None
    if product_errors:
        logprint(f"Smart scanner hittade {len(product_errors)} felaktiga produkter. Se logg och felrapport för detaljer.")
        error_xlsx = export_errors_to_xlsx(product_errors)

    if fallback_used:
        print("⚠️ Exporten gick till CSV istället för XLSX.")
    if error_traceback:
        print("❗ Fullständig felrapport:\n", error_traceback)

    return exported_file, error_xlsx

# ========================
# 8. Run and download (Colab-friendly)
# ========================
def main():
    xlsx_path, error_xlsx_path = enhanced_main_with_scan_and_error_file()
    print("Main scraping/export complete.")
    return xlsx_path, error_xlsx_path

if __name__ == "__main__":
    xlsx_path, error_xlsx_path = enhanced_main_with_scan_and_error_file()
    if xlsx_path:
        print(f"Din fil {xlsx_path} är skapad.")
        try:
            from google.colab import files
            files.download(xlsx_path)
        except ImportError:
            pass
    if error_xlsx_path:
        print(f"Felrapport {error_xlsx_path} är skapad.")
        try:
            from google.colab import files
            files.download(error_xlsx_path)
        except ImportError:
            pass
    if not xlsx_path:
        print("Ingen produktfil skapades.")
    if not error_xlsx_path:
        print("Ingen felrapport skapades.")

